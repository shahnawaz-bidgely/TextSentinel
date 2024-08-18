import os
import re
import nltk
import logging
from nltk.tokenize import word_tokenize, sent_tokenize
from collections import Counter
from openpyxl import load_workbook
import syllapy

nltk.download('punkt', force=True)


directory = './output/extracted/'
output_path = './output/OutputDataStructure.xlsx'
stopwords_path = './input/stop_words/'


positive_words = set(["good", "nice", "great", "awesome", "outstanding", "excellent", "positive", "fortunate", "correct", "superior"])
negative_words = set(["bad", "nasty", "poor", "terrible", "inferior", "wrong", "negative", "unfortunate", "awful", "horrible"])

def get_stopwords(directory):
    stopwords = set()
    for filename in os.listdir(directory):
        if filename.endswith(".txt"):
            filepath = os.path.join(directory, filename)
            try:
                with open(filepath, 'r', encoding='utf-8') as file:
                    words = file.read().splitlines()
            except UnicodeDecodeError:
                logging.warning(f"UTF-8 decoding failed for file: {filename}. Trying ISO-8859-1 encoding.")
                try:
                    with open(filepath, 'r', encoding='iso-8859-1') as file:
                        words = file.read().splitlines()
                except UnicodeDecodeError:
                    logging.error(f"ISO-8859-1 decoding also failed for file: {filename}. Skipping this file.")
                    continue
            stopwords.update(words)
    return stopwords



stop_words = get_stopwords(stopwords_path)

def compute_variables(text):
    sentences = sent_tokenize(text)
    words = word_tokenize(re.sub(r'[^\w\s]', '', text.lower()))

    
    filtered_final_words = [word for word in words if word not in stop_words]

   
    positive_scores = sum(1 for word in filtered_final_words if word in positive_words)
    negative_scores = sum(1 for word in filtered_final_words if word in negative_words)
    word_count = len(filtered_final_words)

    polarity_score = (positive_scores - negative_scores) / ((positive_scores + negative_scores) + 0.000001)
    subjectivity_score = (positive_scores + negative_scores) / (word_count + 0.000001)

    avg_sentence_length = word_count / len(sentences)

    complex_word_count = sum(1 for word in filtered_final_words if syllapy.count(word) > 2)
    syllables_per_word = sum(syllapy.count(word) for word in filtered_final_words) / word_count

    percentage_complex_words = (complex_word_count / word_count) * 100
    fog_index = 0.4 * (avg_sentence_length + percentage_complex_words)

    avg_words_per_sentence = word_count / len(sentences)

    personal_pronouns = re.findall(r'\b(i|we|my|ours|us)\b', text, re.I)
    personal_pronouns_count = len(personal_pronouns)

    avg_word_length = sum(len(word) for word in filtered_final_words) / word_count

    return [
        positive_scores,
        negative_scores,
        polarity_score,
        subjectivity_score,
        avg_sentence_length,
        percentage_complex_words,
        fog_index,
        avg_words_per_sentence,
        complex_word_count,
        word_count,
        syllables_per_word,
        personal_pronouns_count,
        avg_word_length
    ]

wb = load_workbook(output_path)
sheet = wb.active

for filename in os.listdir(directory):
    if filename.endswith(".txt"):
        filepath = os.path.join(directory, filename)
        with open(filepath, 'r') as file:
            text = file.read()

        file_name = os.path.splitext(filename)[0]
        match_found = False
        for row in range(1, sheet.max_row + 1):
            if sheet.cell(row=row, column=1).value == file_name:
                match_found = True
                values = compute_variables(text)

                for i, value in enumerate(values, start=3): 
                    sheet.cell(row=row, column=i, value=value)

                logging.info(f"Successfully processed and saved data for file '{filename}' into row {row}.")
                break
            

        if not match_found:
            logging.warning(f"No matching entry found in Column A for file '{filename}'. Skipping this file.")

wb.save(output_path)
logging.info("All files processed successfully. Workbook saved.")
